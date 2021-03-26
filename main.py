"""
Description: python script for check and retrieving data from xlsx,
and render correct data to another xlsx
Author: Tauassar Tatiyev
"""
import logging
from openpyxl import load_workbook

import log
import settings
from utils import org_form_to_start, write_xlsx, write_incorrect, write_statistics

log.setup_logging('DEBUG' if settings.DEBUG else 'INFO')
logger = logging.getLogger(__name__)


def process_xlsx(name, index):
    duplicates_count = 0
    visited = []
    wb = load_workbook(name)
    sheets = wb.sheetnames
    logger.debug(sheets)
    work_sheet = wb[sheets[index]]
    no_duplicates_list = []
    for value in work_sheet.iter_rows(values_only=True):
        if value[1] not in visited:
            no_duplicates_list.append(value)
            visited.append(value[1])
        else:
            duplicates_count += 1
    corrupted = write_xlsx(no_duplicates_list)
    write_incorrect(corrupted)
    write_statistics(
        len(no_duplicates_list), duplicates_count, len(corrupted))


if __name__ == '__main__':
    process_xlsx(settings.inp_file, 0)
    iin = 'Биг Рэд Интернэшнл КЗ  ТОО'
    logger.debug(org_form_to_start(iin))
