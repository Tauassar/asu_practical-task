import datetime
import logging
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font

from settings import correct_file, incorrect_file, statistics_file, organisational_forms

logger = logging.getLogger(__name__)


def check_iin(iin):
    """
    Function to check whether the input iin is correct or not
    :param iin:str: actual iin
    :return:bool True if iin is correct, false otherwise
    """
    try:
        if len(iin) != 12:
            raise IndexError
        # retrieve last number to check it with control sum
        a12 = int(iin[-1])
        # iin numbers to be used in control sum calculations
        numbers = iin[:-1]
        numbers_sum = 0
        for i in range(len(numbers)):
            numbers_sum += int(numbers[i]) * (i + 1)
        control_num = numbers_sum % 11
        if control_num == 10:
            weights = [3, 4, 5, 6, 7, 8, 9, 10, 11, 1, 2]
            numbers_sum = 0
            for i in range(len(numbers)):
                numbers_sum += int(numbers[i]) * weights[i]
            control_num = numbers_sum % 11
            if 0 > control_num > 9:
                return False
        return True if control_num == a12 else False
    except IndexError:
        logger.warning('Length of IIN/BIN is incorrect')
        return False


def date_check(date):
    """
    Check date in first 4 digits of BIN
    :param date:str
    :return:bool True if not earlier than independence year
    and not later than today's day, false otherwise
    """
    # independence date
    initial_date = '0131'
    independence_date = datetime.datetime.strptime(initial_date, '%m%y')
    current_date = datetime.datetime.now()
    try:
        registration_date = datetime.datetime.strptime(date, '%m%y')
    except ValueError as e:
        logger.warning('Incorrect month for BIN\n {0}'.format(e))
        return False

    if independence_date > registration_date > current_date:
        logger.info('BIN date is incorrect')
        return False
    return True


def check_bin(bin_number):
    """
    Check whether bin is correct or not
    :param bin_number:str actual bin number
    :return:bool True if correct, false otherwise
    """
    # registration date check
    if not date_check(bin_number[0:4]):
        return False
    logger.debug('date check passed')
    if int(bin_number[4]) not in [4, 5, 6]:
        logger.info('second part test failed')
        return False
    logger.debug('second check passed')
    if int(bin_number[5]) not in [0, 1, 2, 3]:
        logger.debug(int(bin_number[5]))
        logger.info('third part test failed')
        return False
    logger.debug('third part test passed')

    # no check for 4th part applied,
    # since there is no data about registration number

    # According to https://adilet.zan.kz/rus/docs/P030000565_
    # bin checked the same way as iin
    return check_iin(bin_number)


def iin_or_bin(number):
    """
    Check if given number is BIN or IIN and sends to appropriate function
    :param number:str BIN/IIN number
    :return:bool True if correct, false otherwise
    """
    try:
        if len(number) != 12:
            logger.warning('Given IIN/BIN is too short: {0}'.format(number))
            return False
        int(number)
        if int(number[4]) in [0, 1, 2, 3]:
            return check_iin(number)
        elif int(number[4]) in [4, 5, 6]:
            return check_iin(number)
        else:
            logger.info(
                '{0} does not match either IIN or BIN pattern'.format(number))
            return False
    except TypeError as e:
        logger.warning('Error in bin_or_iin: {0}\n Number: {1}'.format(e, number))
    except ValueError:
        logger.warning('Given IIN/BIN is not a number: {0}'.format(number))
        return False


def check_single_org_form(org_name, org_form):
    """
    move organisational form to beginning
    :param org_name:str name of the organisation
    :param org_form: TOO ИП etc.
    :return:str converted form
    """
    if re.search(org_form, org_name):
        if re.search(' '+org_form, org_name):
            org_name = re.sub(' ' + org_form, '', org_name)
        else:
            org_name = re.sub(org_form, '', org_name)
        logger.debug("{0} match".format(org_form))
        return '{0} {1}'.format(org_form, org_name)
    logger.debug("{0} no match".format(org_form))
    return None


def org_form_to_start(org_name):
    """
    driver for check_single_org_form,
    sends organisational form to the dedicated function
    :param org_name:str name of organisation
    :return:str proceeded string
    """
    org_forms = organisational_forms
    for form in org_forms:
        out = check_single_org_form(org_name, form)
        if out:
            return out
    # in case organisation has no org_form
    return org_name


# ASSUME every person submitted IIN is 'ФЛ'
# According to https://adilet.zan.kz/rus/docs/P030000565_
# leading to the fact, that everybody with 0, 1, 2, 3, 6 are 'ФЛ'
def fl_or_ul(bin_iin):
    """
    Checks whether organisation belongs to entity or actual person
    :param bin_iin:str BIN number
    :return:str type of organisation
    """
    logger.debug('{0} passed to fl_ul check'.format(bin_iin[4]))
    if int(bin_iin[4]) in [0, 1, 2, 3, 6]:
        return 'ФЛ'
    elif int(bin_iin[4]) in [4, 5]:
        return 'ЮЛ'
    else:
        logger.info('second part of IIN/BIN is out of scope')
        return 'ERROR'


def make_bold(sheet, cells=('A1', 'B1', 'C1')):
    """
    Make upper title cells bold
    :param sheet:object rendered sheet
    """
    for cell_pos in cells:
        cell = sheet[cell_pos]
        cell.font = Font(bold=True)


def save_to_file(sheet, book, a_width, b_width, file_location):
    """
    Actual save function to save rendered sheets to file
    :param sheet:object rendered sheet
    :param book:object xlsx file needed to save
    :param a_width:int column a width
    :param b_width:int column a width
    :param file_location:str location where file needed to save
    """
    try:
        sheet.column_dimensions['A'].width = a_width + 5
        sheet.column_dimensions['B'].width = b_width + 5
        book.save(file_location)
        logger.info(file_location + ' successfully rendered')
    except FileNotFoundError:
        os.mkdir(re.findall("^[a-z, A-z]*/", file_location)[0])
        book.save(file_location)


def write_incorrect(incorrect_list):
    """
    Write data to 'Некорректные БИН'
    :param incorrect_list:list list of data with incorrect BIN/IIN
    """
    book = Workbook()
    sheet = book.active
    a_width = 0
    b_width = 0
    sheet.cell(row=1, column=1).value = 'БИН/ИИН'
    sheet.cell(row=1, column=2).value = 'Контрагент'
    make_bold(sheet)
    for i in range(len(incorrect_list)):
        sheet.cell(row=i + 2, column=1).value = incorrect_list[i][0]
        sheet.cell(row=i + 2, column=2).value = incorrect_list[i][1]
        a = len(str(incorrect_list[i][0]))
        if a > a_width:
            a_width = a
        b = len(incorrect_list[i][1])
        if b > b_width:
            b_width = b
    save_to_file(sheet, book, a_width, b_width, incorrect_file)


def write_xlsx(inp_list):
    """
    Write data to 'Итоговый файл'
    :param inp_list:list list with no duplicates
    :return:list list of data with incorrect IINs
    """
    book = Workbook()
    sheet = book.active
    a_width = 0
    b_width = 0
    incorrect_list = []
    compensation = 2
    heading = inp_list.pop(0)
    sheet.cell(row=1, column=1).value = heading[0]
    sheet.cell(row=1, column=2).value = org_form_to_start(heading[1])
    sheet.cell(row=1, column=3).value = 'ФЛ/ЮЛ'
    make_bold(sheet)
    for i in range(len(inp_list)):
        if not iin_or_bin(str(inp_list[i][0])):
            incorrect_list.append(inp_list[i])
            compensation -= 1
            continue
        sheet.cell(row=i + compensation, column=1).value = inp_list[i][0]
        sheet.cell(row=i + compensation, column=2).value = org_form_to_start(inp_list[i][1])
        sheet.cell(row=i + compensation, column=3).value = fl_or_ul(str(inp_list[i][0]))
        a = len(str(inp_list[i][0]))
        if a > a_width:
            a_width = a
        b = len(inp_list[i][1])
        if b > b_width:
            b_width = b
    save_to_file(sheet, book, a_width, b_width, correct_file)
    return incorrect_list


def write_statistics(filtered, duplcated, incorrect):
    """
    Write data to 'out/Статистика.xlsx'
    :param filtered:int count of unique rows in input
    :param duplcated:int count of non-unique rows in input
    :param incorrect:int count of incorrect BIN/IIN
    """
    book = Workbook()
    sheet = book.active
    total = filtered + duplcated
    a_width = len('Некорректные БИН')
    b_width = 10
    sheet.cell(row=1, column=1).value = 'Наименование'
    sheet.cell(row=1, column=2).value = 'Значение'
    sheet.cell(row=1, column=3).value = 'Процент'
    make_bold(sheet)
    names = ['Тотал', 'Без дупликатов', 'Дупликаты', 'Некорректные БИН']
    values = [total, filtered, duplcated, incorrect]
    for i in range(len(names)):
        sheet.cell(row=i+2, column=1).value = names[i]
        sheet.cell(row=i+2, column=2).value = values[i]
        sheet.cell(row=i+2, column=3).value = '%.2f' % float(values[i] / total * 100)
    sheet.column_dimensions['C'].width = 15
    save_to_file(sheet, book, a_width, b_width, statistics_file)
